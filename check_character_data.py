#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
角色基础属性数据校验脚本

功能：读取「角色基础属性.xlsx」中「角色」工作表，按规则校验数据，
      将问题写入 check_report.xlsx，并在控制台输出汇总信息。
"""

import os
import sys
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# 常量配置
# ---------------------------------------------------------------------------

# 输入文件名（与脚本同目录）
INPUT_FILENAME = "角色基础属性.xlsx"
# 目标工作表名称
SHEET_NAME = "角色"
# 输出报告文件名
OUTPUT_FILENAME = "check_report.xlsx"

# 需要校验数值范围的列及其上下限（含边界）
RANGE_RULES = {
    "生命值": (1000, 50000),
    "攻击力": (50, 5000),
    "防御力": (50, 5000),
}

# 「星级」列允许的值
ALLOWED_STAR_VALUES = {4, 5}

# 报告表头
REPORT_HEADERS = ["行号", "字段", "问题描述"]


def get_script_directory() -> str:
    """获取脚本所在目录的绝对路径。"""
    return os.path.dirname(os.path.abspath(__file__))


def load_character_sheet(filepath: str):
    """
    加载 Excel 文件并返回「角色」工作表对象。

    :param filepath: Excel 文件的完整路径
    :return: openpyxl 工作表对象
    :raises FileNotFoundError: 文件不存在
    :raises KeyError: 找不到名为「角色」的工作表
    """
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"找不到文件：{filepath}")

    workbook = load_workbook(filepath, data_only=True)

    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(
            f"工作簿中不存在名为「{SHEET_NAME}」的工作表。"
            f"现有工作表：{workbook.sheetnames}"
        )

    return workbook[SHEET_NAME]


def build_column_index(header_row) -> dict:
    """
    根据表头行构建「列名 -> 列索引」映射。

    :param header_row: 工作表第一行（表头）
    :return: 字典，键为列名，值为 1-based 列号
    """
    column_index = {}
    for cell in header_row:
        if cell.value is not None:
            column_index[str(cell.value).strip()] = cell.column
    return column_index


def is_empty(value) -> bool:
    """
    判断单元格值是否为空。

    空值包括：None、空字符串、仅含空白字符的字符串。
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def normalize_numeric(value):
    """
    尝试将值转换为数值类型，便于范围校验。

    :return: (是否成功, 数值或原值)
    """
    if is_empty(value):
        return False, value
    try:
        return True, float(value)
    except (TypeError, ValueError):
        return False, value


def add_issue(issues: list, row_num: int, field: str, description: str) -> None:
    """
    向问题列表追加一条记录。

    :param issues: 问题列表
    :param row_num: Excel 行号（1-based，含表头）
    :param field: 出问题的字段名
    :param description: 问题描述
    """
    issues.append({"行号": row_num, "字段": field, "问题描述": description})


def check_duplicate_ids(sheet, col_idx: int, issues: list) -> None:
    """
    规则 a：检查「角色ID」列是否存在重复值。

    同一 ID 出现多次时，为每一次出现都记录一条问题。
    """
    field = "角色ID"
    id_rows = defaultdict(list)

    for row_num in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_num, column=col_idx).value
        if not is_empty(value):
            id_rows[value].append(row_num)

    for role_id, rows in id_rows.items():
        if len(rows) > 1:
            row_list = "、".join(str(r) for r in rows)
            for row_num in rows:
                add_issue(
                    issues,
                    row_num,
                    field,
                    f"角色ID「{role_id}」重复，出现在第 {row_list} 行",
                )


def check_numeric_range(sheet, col_idx: int, field: str, min_val: float, max_val: float, issues: list) -> None:
    """
    规则 b/c/d：检查指定数值列是否在允许范围内。

    空值由空值检查规则单独处理，此处跳过。
    非数值类型会单独报错。
    """
    for row_num in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_num, column=col_idx).value
        if is_empty(value):
            continue

        ok, numeric = normalize_numeric(value)
        if not ok:
            add_issue(
                issues,
                row_num,
                field,
                f"值「{value}」不是有效数值，无法校验范围 [{min_val}, {max_val}]",
            )
            continue

        if numeric < min_val or numeric > max_val:
            add_issue(
                issues,
                row_num,
                field,
                f"值 {numeric} 超出允许范围 [{min_val}, {max_val}]",
            )


def check_empty_values(sheet, column_index: dict, issues: list) -> None:
    """
    规则 e：检查所有列是否存在空值。
    """
    for row_num in range(2, sheet.max_row + 1):
        for field, col_idx in column_index.items():
            value = sheet.cell(row=row_num, column=col_idx).value
            if is_empty(value):
                add_issue(issues, row_num, field, "存在空值")


def check_star_values(sheet, col_idx: int, issues: list) -> None:
    """
    规则 f：检查「星级」列是否只包含 4 或 5。

    空值由空值检查规则处理；非整数或非法值单独报错。
    """
    field = "星级"

    for row_num in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_num, column=col_idx).value
        if is_empty(value):
            continue

        ok, numeric = normalize_numeric(value)
        if not ok:
            add_issue(
                issues,
                row_num,
                field,
                f"值「{value}」不是有效数值，星级只能为 4 或 5",
            )
            continue

        # 星级应为整数
        if numeric != int(numeric):
            add_issue(
                issues,
                row_num,
                field,
                f"值 {numeric} 不是整数，星级只能为 4 或 5",
            )
            continue

        star = int(numeric)
        if star not in ALLOWED_STAR_VALUES:
            add_issue(
                issues,
                row_num,
                field,
                f"值 {star} 非法，星级只能为 4 或 5",
            )


def validate_sheet(sheet) -> list:
    """
    对工作表执行全部校验规则，返回问题列表。

    :return: 问题字典列表，每项含「行号」「字段」「问题描述」
    """
    issues = []

    if sheet.max_row < 2:
        add_issue(issues, 1, "全局", "工作表无数据行（仅有表头或为空）")
        return issues

    column_index = build_column_index(sheet[1])

    # 校验所需列是否都存在
    required_columns = ["角色ID"] + list(RANGE_RULES.keys()) + ["星级"]
    missing_columns = [col for col in required_columns if col not in column_index]
    if missing_columns:
        for col in missing_columns:
            add_issue(issues, 1, "全局", f"缺少必需列「{col}」")
        return issues

    # 规则 a：角色ID 重复
    check_duplicate_ids(sheet, column_index["角色ID"], issues)

    # 规则 b/c/d：数值范围
    for field, (min_val, max_val) in RANGE_RULES.items():
        check_numeric_range(
            sheet, column_index[field], field, min_val, max_val, issues
        )

    # 规则 e：空值
    check_empty_values(sheet, column_index, issues)

    # 规则 f：星级
    check_star_values(sheet, column_index["星级"], issues)

    return issues


def write_report(issues: list, output_path: str) -> None:
    """
    将问题列表写入 Excel 报告文件。

    :param issues: 问题列表
    :param output_path: 输出文件完整路径
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "校验报告"

    # 写入表头并设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(REPORT_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入问题数据
    for row_num, issue in enumerate(issues, start=2):
        sheet.cell(row=row_num, column=1, value=issue["行号"])
        sheet.cell(row=row_num, column=2, value=issue["字段"])
        sheet.cell(row=row_num, column=3, value=issue["问题描述"])

    # 自动调整列宽（简单估算）
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 60

    workbook.save(output_path)


def print_summary(issues: list, data_row_count: int) -> None:
    """
    在控制台输出校验汇总信息。

    :param issues: 问题列表
    :param data_row_count: 数据行数量（不含表头）
    """
    print("=" * 60)
    print("角色基础属性数据校验报告")
    print("=" * 60)
    print(f"校验数据行数：{data_row_count}")
    print(f"发现问题总数：{len(issues)}")

    if not issues:
        print("\n✓ 全部校验通过，未发现问题。")
        return

    # 按字段统计问题数量
    field_stats = defaultdict(int)
    for issue in issues:
        field_stats[issue["字段"]] += 1

    print("\n按字段统计：")
    for field, count in sorted(field_stats.items(), key=lambda x: (-x[1], x[0])):
        print(f"  - {field}：{count} 条")

    print("\n问题明细（前 10 条）：")
    for issue in issues[:10]:
        print(
            f"  第 {issue['行号']} 行 | {issue['字段']} | {issue['问题描述']}"
        )
    if len(issues) > 10:
        print(f"  ... 其余 {len(issues) - 10} 条请查看 {OUTPUT_FILENAME}")

    print("=" * 60)


def main() -> int:
    """
    主入口：加载数据、执行校验、写报告、输出汇总。

    :return: 进程退出码，0 表示无问题，1 表示有问题或运行失败
    """
    script_dir = get_script_directory()
    input_path = os.path.join(script_dir, INPUT_FILENAME)
    output_path = os.path.join(script_dir, OUTPUT_FILENAME)

    try:
        print(f"正在读取：{input_path}")
        sheet = load_character_sheet(input_path)
        data_row_count = max(0, sheet.max_row - 1)

        print("正在执行数据校验...")
        issues = validate_sheet(sheet)

        print(f"正在写入报告：{output_path}")
        write_report(issues, output_path)

        print_summary(issues, data_row_count)

        return 0 if not issues else 1

    except FileNotFoundError as exc:
        print(f"[错误] {exc}", file=sys.stderr)
        print("请确认「角色基础属性.xlsx」与脚本位于同一目录。", file=sys.stderr)
        return 1

    except KeyError as exc:
        print(f"[错误] {exc}", file=sys.stderr)
        return 1

    except PermissionError:
        print(
            f"[错误] 无法访问文件，请关闭已打开的 Excel 后重试。",
            file=sys.stderr,
        )
        return 1

    except Exception as exc:
        print(f"[错误] 校验过程中发生异常：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
